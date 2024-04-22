import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import openpyxl  # Añadido openpyxl
import win32com.client as win32

def procesar_archivos(ccpp_file, eess_file, iiee_file):
    
    ccpp_data = pd.read_excel(ccpp_file)
    eess_data = pd.read_excel(eess_file)
    iiee_data = pd.read_excel(iiee_file)
         
    ccpp_huanuco = ccpp_data[ccpp_data['Departamento'] == "HUANUCO"]
    eess_huanuco = eess_data[eess_data['Departamento'] == "HUANUCO"]
    iiee_huanuco = iiee_data[iiee_data['Departamento'] == "HUANUCO"]

    provincias_huanuco = ["AMBO", "DOS DE MAYO", "HUACAYBAMBA", "HUAMALIES", "LEONCIO PRADO", 
                          "MARAÑON", "PACHITEA", "PUERTO INCA", "LAURICOCHA", "YAROWILCA", "HUANUCO"]

    for provincia in provincias_huanuco:
        os.makedirs(f"Inundaciones/{provincia}", exist_ok=True)
        os.makedirs(f"Movimiento_en_masa/{provincia}", exist_ok=True)

        ccpp_provincia = ccpp_huanuco[ccpp_huanuco['Provincia'] == provincia].copy()
        eess_provincia = eess_huanuco[eess_huanuco['Provincia'] == provincia].copy()
        iiee_provincia = iiee_huanuco[iiee_huanuco['Provincia'] == provincia].copy()

        masa_columns_ccpp = [col for col in ccpp_provincia.columns if 'masa' in col.lower()]
        masa_columns_eess = [col for col in eess_provincia.columns if 'masa' in col.lower()]
        masa_columns_iiee = [col for col in iiee_provincia.columns if 'masa' in col.lower()]

        ccpp_provincia.drop(columns=masa_columns_ccpp, inplace=True)
        eess_provincia.drop(columns=masa_columns_eess, inplace=True)
        iiee_provincia.drop(columns=masa_columns_iiee, inplace=True)

        ccpp_provincia.to_excel(f"Inundaciones/{provincia}/CCPP_HUANUCO_{provincia}.xlsx", index=False)
        eess_provincia.to_excel(f"Inundaciones/{provincia}/EESS_HUANUCO_{provincia}.xlsx", index=False)
        iiee_provincia.to_excel(f"Inundaciones/{provincia}/IIEE_HUANUCO_{provincia}.xlsx", index=False)

        ccpp_provincia = ccpp_huanuco[ccpp_huanuco['Provincia'] == provincia].copy()
        eess_provincia = eess_huanuco[eess_huanuco['Provincia'] == provincia].copy()
        iiee_provincia = iiee_huanuco[iiee_huanuco['Provincia'] == provincia].copy()

        inundaciones_columns_ccpp = [col for col in ccpp_provincia.columns if 'inundaciones' in col.lower()]
        inundaciones_columns_eess = [col for col in eess_provincia.columns if 'inundaciones' in col.lower()]
        inundaciones_columns_iiee = [col for col in iiee_provincia.columns if 'inundaciones' in col.lower()]

        ccpp_provincia.drop(columns=inundaciones_columns_ccpp, inplace=True)
        eess_provincia.drop(columns=inundaciones_columns_eess, inplace=True)
        iiee_provincia.drop(columns=inundaciones_columns_iiee, inplace=True)

        ccpp_provincia.to_excel(f"Movimiento_en_masa/{provincia}/CCPP_HUANUCO_{provincia}.xlsx", index=False)
        eess_provincia.to_excel(f"Movimiento_en_masa/{provincia}/EESS_HUANUCO_{provincia}.xlsx", index=False)
        iiee_provincia.to_excel(f"Movimiento_en_masa/{provincia}/IIEE_HUANUCO_{provincia}.xlsx", index=False)

    ccpp_huanuco.to_excel("Inundaciones/CCPP_HUANUCO.xlsx", index=False)
    eess_huanuco.to_excel("Inundaciones/EESS_HUANUCO.xlsx", index=False)
    iiee_huanuco.to_excel("Inundaciones/IIEE_HUANUCO.xlsx", index=False)

    ccpp_huanuco.to_excel("Movimiento_en_masa/CCPP_HUANUCO.xlsx", index=False)
    eess_huanuco.to_excel("Movimiento_en_masa/EESS_HUANUCO.xlsx", index=False)
    iiee_huanuco.to_excel("Movimiento_en_masa/IIEE_HUANUCO.xlsx", index=False)

    convertir_todos_a_tablas("Inundaciones")
    convertir_todos_a_tablas("Movimiento_en_masa")

def seleccionar_archivos():
    ccpp_file = filedialog.askopenfilename(title="Seleccionar archivo de Centros Poblados")
    eess_file = filedialog.askopenfilename(title="Seleccionar archivo de Establecimientos de Salud")
    iiee_file = filedialog.askopenfilename(title="Seleccionar archivo de Instituciones Educativas")

    if ccpp_file and eess_file and iiee_file:
        procesar_archivos(ccpp_file, eess_file, iiee_file)
        messagebox.showinfo("Proceso completado", "Los archivos han sido procesados correctamente.")

def convertir_todos_a_tablas(carpetas):
    for root, dirs, files in os.walk(carpetas):
        for file in files:
            if file.endswith(".xlsx"):
                convertir_a_tabla(os.path.join(root, file))

def convertir_a_tabla(file_path):
    
    wb = load_workbook(file_path)
    ws = wb.active

    
    tbl = pd.read_excel(file_path)

    for r_idx, row in enumerate(dataframe_to_rows(tbl, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    
    for cell in ws["1"]:
        cell.fill = PatternFill(start_color="38e3ff", end_color="38e3ff", fill_type="solid")

    
    max_row = ws.max_row
    max_col = ws.max_column
    tbl_range = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

    # Crear la tabla en el rango definido
    tbl = openpyxl.worksheet.table.Table(displayName="Tabla1", ref=tbl_range)

    # Agregar la tabla a la hoja de cálculo
    ws.add_table(tbl)

    # Guardar los cambios en el archivo Excel
    wb.save(file_path)

# Crear la ventana principal
root = tk.Tk()
root.title("Procesador de Archivos Excel")

# Botón para seleccionar archivos
button = tk.Button(root, text="Seleccionar archivos Excel", command=seleccionar_archivos)
button.pack(pady=20)

# Ejecutar la aplicación
root.mainloop()
